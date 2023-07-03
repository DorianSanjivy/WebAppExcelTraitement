from flask import Flask, request, send_file, render_template, jsonify
from flask_uploads import UploadSet, configure_uploads, DOCUMENTS
import pandas as pd
import requests
import os
import webbrowser
import threading
from openpyxl.drawing.image import Image
import time
from matplotlib import pyplot as plt
from openpyxl import load_workbook
import uuid

app = Flask(__name__)

# Configuration for file uploads
documents = UploadSet('documents', DOCUMENTS)
app.config['UPLOADED_DOCUMENTS_DEST'] = 'uploads'
configure_uploads(app, documents)

# Configuration for uploading the file
@app.route('/upload', methods=['GET', 'POST'])
def upload():
   if request.method == 'POST' and 'document' in request.files and 'language' in request.form:
       filename = documents.save(request.files['document'])
       language = request.form.get('language')
       if language == 'en':
           new_filename, sentiment_image_path, emotion_image_path = process_file_en(os.path.join('uploads', filename))
           return jsonify({'filename': new_filename, 'sentiment_image': sentiment_image_path, 'emotion_image': emotion_image_path})
       elif language == 'fr':
           new_filename, sentiment_image_path = process_file_fr(os.path.join('uploads', filename))
           return jsonify({'filename': new_filename, 'sentiment_image': sentiment_image_path})
   return render_template('upload.html')

# Configuration for downloading the file
@app.route('/download/<filename>')
def download(filename):
    return send_file(os.path.join('uploads', filename), as_attachment=True)

@app.route('/images/<filename>')
def serve_image(filename):
    return send_file(os.path.join('static', filename), mimetype='image/png')

# Function to process the file
def process_file_en(filepath):
    time.sleep(1)  # Wait a bit before processing the file
    df = pd.read_excel(filepath)

    if 'Sentiment' not in df.columns:
        df['Sentiment'] = ''
    if 'Emotion' not in df.columns:
        df['Emotion'] = ''
    if 'Summary' not in df.columns:
        df['Summary'] = ''

    #API URL
    url = "https://testapi5-rvalk7bhha-od.a.run.app/predict"

    #for each review in the dataframe, send a request to the API
    for index, row in df.iterrows():
        sentence = row['Review']
        data = {'sentence': sentence}
        response = requests.post(url, json=data)

        if response.status_code == 200:
            result = response.json()
            df.at[index, 'Sentiment'] = result['sentiment']
            df.at[index, 'Emotion'] = result['emotion']
            df.at[index, 'Summary'] = result['summary']

    base, extension = os.path.splitext(filepath)
    new_filepath = f"{base}_processed{extension}"

    df.to_excel(new_filepath, index=False)

    sentiment_counts = df['Sentiment'].value_counts()
    emotion_counts = df['Emotion'].value_counts()

    # Save the sentiment and emotion counts in the excel file
    with pd.ExcelWriter(new_filepath, mode='a', engine='openpyxl') as writer:
        sentiment_counts.to_frame().to_excel(writer, sheet_name='Sentiment Counts')
        emotion_counts.to_frame().to_excel(writer, sheet_name='Emotion Counts')

    # Avant de sauvegarder les images, assurez-vous que le dossier 'static' existe
    if not os.path.exists('static'):
        os.makedirs('static')

    # Generate unique id for the images
    uid = uuid.uuid4()

    # Generate the images
    plt.figure(figsize=(5, 5))
    sentiment_counts.plot(kind='pie', autopct='%1.1f%%')
    plt.title("Sentiment")
    plt.ylabel("")
    sentiment_image_path = f'static/sentiment_pie_{uid}.png'
    plt.savefig(sentiment_image_path)

    plt.figure(figsize=(5, 5))
    emotion_counts.plot(kind='pie', autopct='%1.1f%%')
    plt.title("Emotion")
    plt.ylabel("")
    emotion_image_path = f'static/emotion_pie_{uid}.png'
    plt.savefig(emotion_image_path)

    book = load_workbook(new_filepath)

    # Add the images to the excel file
    sentiment_img = Image(sentiment_image_path)
    sentiment_img.width = 400
    sentiment_img.height = 400
    book['Sentiment Counts'].add_image(sentiment_img, 'A1')

    emotion_img = Image(emotion_image_path)
    emotion_img.width = 400
    emotion_img.height = 400
    book['Emotion Counts'].add_image(emotion_img, 'A1')

    book.save(new_filepath)  # Save the workbook after adding the images

    return os.path.basename(new_filepath), sentiment_image_path, emotion_image_path  # Return the new filename and the paths to the images




# Function to process the file (French)
def process_file_fr(filepath):
    time.sleep(1)  # Wait a bit before processing the file
    df = pd.read_excel(filepath)

    if 'Sentiment' not in df.columns:
        df['Sentiment'] = ''

    #API URL
    url = "https://testapi6-rvalk7bhha-od.a.run.app/predict"

    #for each review in the dataframe, send a request to the API
    for index, row in df.iterrows():
        sentence = row['Review']
        data = {'sentence': sentence}
        response = requests.post(url, json=data)

        if response.status_code == 200:
            result = response.json()
            df.at[index, 'Sentiment'] = result['sentiment']

    base, extension = os.path.splitext(filepath)
    new_filepath = f"{base}_processed{extension}"

    df.to_excel(new_filepath, index=False)

    sentiment_counts = df['Sentiment'].value_counts()

    # Save the sentiment counts in the excel file
    with pd.ExcelWriter(new_filepath, mode='a', engine='openpyxl') as writer:
        sentiment_counts.to_frame().to_excel(writer, sheet_name='Sentiment Counts')

    # Generate unique id for the image
    uid = uuid.uuid4()

    # Avant de sauvegarder les images, assurez-vous que le dossier 'static' existe
    if not os.path.exists('static'):
        os.makedirs('static')

    # Generate the image
    plt.figure(figsize=(5, 5))
    sentiment_counts.plot(kind='pie', autopct='%1.1f%%')
    plt.title("Sentiment")
    plt.ylabel("")
    sentiment_image_path = f'static/sentiment_pie_{uid}.png'
    plt.savefig(sentiment_image_path)

    book = load_workbook(new_filepath)

    # Add the images to the excel file
    sentiment_img = Image(sentiment_image_path)
    sentiment_img.width = 400
    sentiment_img.height = 400
    book['Sentiment Counts'].add_image(sentiment_img, 'A1')

    book.save(new_filepath)  # Save the workbook after adding the images

    # Return the new filename and the sentiment image path
    return os.path.basename(new_filepath), sentiment_image_path

def open_browser():
    webbrowser.open_new('http://localhost:5000/upload')

if __name__ == '__main__':
    threading.Timer(1, open_browser).start()
    app.run(port=5000)

